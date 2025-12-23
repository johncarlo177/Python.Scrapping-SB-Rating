import re
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

ua = UserAgent()
USER_AGENT = ua.random
ChromeDriverPath = "C:/chromedriver/chromedriver.exe"

BASE_URL = 'https://www.sportsbet.com.au'
FILE_NAME = 'Race Meetings.xlsm'
target_column = 23
ALLOWED_MEETINGS = ['(VIC)', '(NSW)', '(QLD)', '(SA)', '(WA)', '(NT)', '(TAS)', '(ACT)', '(NZ)', '(NZL)']
FS = {}
SR = {}

def setup_driver():
    options = Options()
    options.headless = True
    options.add_argument("--disable-images")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")
    options.add_argument(f"--user-agent={USER_AGENT}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--no-first-run")
    options.add_argument("--disable-site-isolation-trials")

    service = Service(ChromeDriverPath)
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(800)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver

def extract_sb_rating(driver, race_url, sheet_name):
    global SR

    driver.get(BASE_URL + race_url)
    wait = WebDriverWait(driver, 20)

    # 1️⃣ Expand Form
    try:
        expand_btn = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "span[data-automation-id='racecard-expand-form']")
            )
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", expand_btn)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", expand_btn)
    except Exception:
        pass  # already expanded

    # 2️⃣ Wait for shortforms
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "div[data-automation-id^='shortform-']")
        )
    )

    soup = BeautifulSoup(driver.page_source, "html.parser")

    shortforms = soup.select("div[data-automation-id^='shortform-']")
    print(f"🧩 Shortforms found: {len(shortforms)}")

    for sf in shortforms:
        sf_id = sf.get("data-automation-id")

        # Extract runner ID
        match = re.search(r"shortform-(\d+)", sf_id)
        if not match:
            continue

        runner_id = match.group(1)

        # 🔑 Find matching racecard
        racecard = soup.select_one(
            f"div[data-automation-id='racecard-outcome-{runner_id}']"
        )
        if not racecard:
            continue

        # Horse name
        name_el = racecard.select_one(
            "div[data-automation-id='racecard-outcome-name'] span"
        )
        if not name_el:
            continue

        horse_name = re.sub(r"^\d+\.\s*", "", name_el.get_text(strip=True))

        # SB Rating
        sb_el = sf.select_one(
            "div[data-automation-id='shortform-SB Rating'] span:last-child"
        )
        if not sb_el:
            continue

        sb_rating = sb_el.get_text(strip=True)

        SR.setdefault(sheet_name, {})
        SR[sheet_name][horse_name] = sb_rating

        print(f"✅ {sheet_name} | {horse_name} → SB Rating {sb_rating}")


def disable_international_filter(driver):
    wait = WebDriverWait(driver, 20)

    try:
        # Detect ON state
        on_state = driver.find_elements(
            By.CSS_SELECTOR,
            "div[data-automation-id='filter-button-international-on']"
        )

        if not on_state:
            print("Int'l already OFF")
            return

        print("Int'l filter ON → disabling")

        # Click the LABEL (this is critical)
        label = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "label[for='ALL_RACING_PAGEINTERNATIONAL']")
            )
        )

        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", label
        )
        time.sleep(0.3)

        driver.execute_script("arguments[0].click();", label)

        # WAIT FOR OFF STATE
        wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div[data-automation-id='filter-button-international-off']")
            )
        )

        # WAIT FOR TABLE RE-DRAW (critical)
        wait.until(
            EC.staleness_of(
                driver.find_element(
                    By.CSS_SELECTOR,
                    "td[data-automation-id^='horse-racing-section-row-']"
                )
            )
        )

        print("Int'l filter OFF")

    except Exception as e:
        print("Failed to disable Int'l filter:", e)


def get_races_for_meeting(driver, excel_meeting_name):
    wait = WebDriverWait(driver, 20)

    driver.get(BASE_URL + "/racing-schedule")

    # Wait initial table
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "td[data-automation-id$='-meeting-cell']")
        )
    )

    # Disable Int'l
    disable_international_filter(driver)

    # Re-wait after filter
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "td[data-automation-id$='-meeting-cell']")
        )
    )

    soup = BeautifulSoup(driver.page_source, "html.parser")

    target_meeting = normalize_meeting(excel_meeting_name)

    race_links = []

    # 1️⃣ Find ALL meeting cells
    meeting_cells = soup.select(
        "td[data-automation-id^='horse-racing-section-row-'][data-automation-id$='-meeting-cell']"
    )

    for meeting_cell in meeting_cells:
        name_el = meeting_cell.select_one(
            "span[data-automation-id$='-meeting-name']"
        )

        if not name_el:
            continue

        meeting_text = name_el.get_text(strip=True)

        if normalize_meeting(meeting_text) != target_meeting:
            continue

        # ✅ row number
        row_match = re.search(
            r"row-(\d+)-meeting-cell",
            meeting_cell["data-automation-id"]
        )
        if not row_match:
            continue

        row_number = row_match.group(1)
        print(f"🎯 Matched meeting: {meeting_text} (row {row_number})")

        # ✅ find race cells for that row
        race_cells = soup.select(
            f"td[data-automation-id^='horse-racing-section-row-{row_number}-col-'][data-automation-id$='-event-cell']"
        )

        for td in race_cells:
            a = td.find("a", href=True)
            if a:
                race_links.append(a["href"])

        break  # only one meeting

    print(f"🏇 Races found for {excel_meeting_name}: {len(race_links)}")
    return race_links


def get_meetings_from_excel():
    workbook = load_workbook(FILE_NAME, keep_vba=True)

    meetings = []  # [(sheet_name, meeting_name)]

    for sheet in workbook.worksheets:
        meeting = sheet["G1"].value
        if meeting:
            meetings.append((sheet.title, str(meeting).strip()))

    print("📄 Meetings loaded from Excel:")
    for s, m in meetings:
        print(f"   {s} → {m}")

    return meetings

def normalize_meeting(name: str) -> str:
    return (
        name.strip()
        .lower()
        .replace("(australia)", "")
        .replace("(nz)", "")
    )



def normalize_horse(name: str) -> str:
    return (
        name.strip()
        .upper()          # case-insensitive
        .replace(".", "") # remove dots like "5. "
    )

def save_sb_to_excel(excel_file, SR):
    workbook = load_workbook(excel_file, keep_vba=True)

    for sheet_name, horses in SR.items():
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=1):
            horse_cell = row[3]  # Column D
            if not horse_cell.value:
                continue

            excel_horse = normalize_horse(str(horse_cell.value))

            for sb_horse, sb_rating in horses.items():
                if normalize_horse(sb_horse) == excel_horse:
                    sheet.cell(row=horse_cell.row, column=25, value=sb_rating)
                    break

    workbook.save(excel_file)



def main():
    driver = setup_driver()

    meetings = get_meetings_from_excel()

    for sheet_name, meeting_name in meetings:
        print(f"\n📍 Processing {sheet_name} | Meeting: {meeting_name}")

        race_links = get_races_for_meeting(driver, meeting_name)

        for race_url in race_links:
            extract_sb_rating(driver, race_url, sheet_name)

    driver.quit()
    save_sb_to_excel(FILE_NAME, SR)



if __name__ == '__main__':
    main()
